"""
API router for the Rankings feature.
Handles fund CRUD, uploads, ranking calculations, and exports.
"""
import io
import csv
import uuid
import tempfile
from datetime import date, datetime
from typing import Optional, List

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, desc, asc

from database import get_db
from ranking_models import Fund, FundCategory, FundScore
from ranking_schemas import (
    CategoryOut, FundCreate, FundUpdate, FundOut, FundWithScores,
    FundDetailOut, ScoreOut, RankedFundOut, RankingsResponse,
    CategorySummaryOut, UploadSummary, RecalculateResponse
)
from ranking_calculator import RankingCalculator

router = APIRouter(prefix="/api/rankings", tags=["rankings"])
calculator = RankingCalculator()


# =====================================================================
# HELPER FUNCTIONS
# =====================================================================

def _fund_to_calc_dict(fund: Fund, parent_category: Optional[str] = None) -> dict:
    """Convert a Fund ORM object to a dict for the calculator."""
    d = {}
    for col in Fund.__table__.columns:
        val = getattr(fund, col.name, None)
        d[col.name] = float(val) if val is not None and col.name not in (
            'id', 'ticker', 'name', 'fund_type', 'category_id',
            'inception_date', 'data_as_of_date', 'created_at', 'updated_at'
        ) else val
    d['parent_category'] = parent_category
    return d


def _recalculate_all(db: Session) -> int:
    """Recalculate all fund scores and rankings. Returns count of funds processed."""
    funds = db.query(Fund).all()
    if not funds:
        return 0

    all_scores = []
    for fund in funds:
        cat = db.query(FundCategory).filter(FundCategory.id == fund.category_id).first()
        parent_cat = cat.parent_category if cat else None
        fund_dict = _fund_to_calc_dict(fund, parent_cat)
        scores = calculator.calculate_all_scores(fund_dict)
        scores['fund_id'] = fund.id
        scores['category_id'] = fund.category_id
        all_scores.append(scores)

    # Rank
    calculator.rank_funds(all_scores)

    # Upsert scores
    for score_data in all_scores:
        fund_id = score_data.pop('fund_id')
        score_data.pop('category_id', None)

        existing = db.query(FundScore).filter(FundScore.fund_id == fund_id).first()
        if existing:
            for key, value in score_data.items():
                setattr(existing, key, value)
            existing.calculated_at = datetime.utcnow()
        else:
            score = FundScore(fund_id=fund_id, **score_data)
            db.add(score)

    db.commit()
    return len(funds)


# Column mapping for uploads
COLUMN_MAP = {
    "ticker": "ticker",
    "name": "name",
    "fund type": "fund_type",
    "category": "_category_name",
    "inception date": "inception_date",
    "fund age (years)": "fund_age_years",
    "fund age": "fund_age_years",
    "net expense ratio": "net_expense_ratio",
    "turnover": "turnover",
    "market cap (m)": "market_cap",
    "market cap": "market_cap",
    "yield": "yield_pct",
    "p/e ratio": "pe_ratio",
    "pe ratio": "pe_ratio",
    "p/b ratio": "pb_ratio",
    "pb ratio": "pb_ratio",
    "beta 3yr": "beta_3yr",
    "beta 5yr": "beta_5yr",
    "r-squared 3yr": "r_squared_3yr",
    "r-squared 5yr": "r_squared_5yr",
    "r squared 3yr": "r_squared_3yr",
    "r squared 5yr": "r_squared_5yr",
    "up capture 3yr": "up_capture_3yr",
    "up capture 5yr": "up_capture_5yr",
    "down capture 3yr": "down_capture_3yr",
    "down capture 5yr": "down_capture_5yr",
    "sharpe 3yr": "sharpe_ratio_3yr",
    "sharpe 5yr": "sharpe_ratio_5yr",
    "tracking error 3yr": "tracking_error_3yr",
    "tracking error 5yr": "tracking_error_5yr",
    "sortino 3yr": "sortino_ratio_3yr",
    "sortino 5yr": "sortino_ratio_5yr",
    "treynor 3yr": "treynor_ratio_3yr",
    "treynor 5yr": "treynor_ratio_5yr",
    "info ratio 3yr": "information_ratio_3yr",
    "info ratio 5yr": "information_ratio_5yr",
    "kurtosis 3yr": "kurtosis_3yr",
    "kurtosis 5yr": "kurtosis_5yr",
    "max drawdown 3yr": "max_drawdown_3yr",
    "max drawdown 5yr": "max_drawdown_5yr",
    "skewness 3yr": "skewness_3yr",
    "skewness 5yr": "skewness_5yr",
    "alpha 3yr": "alpha_3yr",
    "alpha 5yr": "alpha_5yr",
    "return qtd": "return_qtd",
    "return ytd": "return_ytd",
    "return 1yr": "return_1yr",
    "return 3yr": "return_3yr",
    "return 5yr": "return_5yr",
    "return 10yr": "return_10yr",
    "bm return 1yr": "bm_return_1yr",
    "bm return 3yr": "bm_return_3yr",
    "bm return 5yr": "bm_return_5yr",
    "bm return 10yr": "bm_return_10yr",
    "batting avg 3yr": "batting_avg_3yr",
    "batting avg 5yr": "batting_avg_5yr",
}

NUMERIC_FIELDS = {
    "fund_age_years", "net_expense_ratio", "turnover", "market_cap", "yield_pct",
    "pe_ratio", "pb_ratio",
    "beta_3yr", "r_squared_3yr", "up_capture_3yr", "down_capture_3yr",
    "sharpe_ratio_3yr", "tracking_error_3yr", "sortino_ratio_3yr", "treynor_ratio_3yr",
    "information_ratio_3yr", "kurtosis_3yr", "max_drawdown_3yr", "skewness_3yr", "alpha_3yr",
    "beta_5yr", "r_squared_5yr", "up_capture_5yr", "down_capture_5yr",
    "sharpe_ratio_5yr", "tracking_error_5yr", "sortino_ratio_5yr", "treynor_ratio_5yr",
    "information_ratio_5yr", "kurtosis_5yr", "max_drawdown_5yr", "skewness_5yr", "alpha_5yr",
    "return_qtd", "return_ytd", "return_1yr", "return_3yr", "return_5yr", "return_10yr",
    "bm_return_1yr", "bm_return_3yr", "bm_return_5yr", "bm_return_10yr",
    "batting_avg_3yr", "batting_avg_5yr",
}


# =====================================================================
# CATEGORIES
# =====================================================================

@router.get("/categories", response_model=List[CategoryOut])
async def list_categories(db: Session = Depends(get_db)):
    """List all fund categories with fund counts."""
    cats = db.query(FundCategory).order_by(FundCategory.display_order).all()
    result = []
    for cat in cats:
        count = db.query(Fund).filter(Fund.category_id == cat.id).count()
        result.append(CategoryOut(
            id=cat.id,
            name=cat.name,
            parent_category=cat.parent_category,
            display_order=cat.display_order,
            fund_count=count
        ))
    return result


# =====================================================================
# FUND CRUD
# =====================================================================

@router.get("/funds", response_model=List[FundWithScores])
async def list_funds(
    category: Optional[int] = None,
    search: Optional[str] = None,
    page: int = 1,
    limit: int = 25,
    db: Session = Depends(get_db)
):
    """List all funds with optional category filter and search."""
    query = db.query(Fund)

    if category:
        query = query.filter(Fund.category_id == category)
    if search:
        search_term = f"%{search}%"
        query = query.filter(
            (Fund.ticker.ilike(search_term)) | (Fund.name.ilike(search_term))
        )

    total = query.count()
    funds = query.offset((page - 1) * limit).limit(limit).all()

    result = []
    for f in funds:
        cat = db.query(FundCategory).filter(FundCategory.id == f.category_id).first()
        scores = db.query(FundScore).filter(FundScore.fund_id == f.id).first()

        fund_out = FundWithScores(
            id=f.id,
            ticker=f.ticker,
            name=f.name,
            fund_type=f.fund_type,
            category_id=f.category_id,
            category_name=cat.name if cat else None,
            parent_category=cat.parent_category if cat else None,
            fund_age_years=float(f.fund_age_years) if f.fund_age_years else None,
            net_expense_ratio=float(f.net_expense_ratio) if f.net_expense_ratio else None,
            turnover=float(f.turnover) if f.turnover else None,
            market_cap=float(f.market_cap) if f.market_cap else None,
            yield_pct=float(f.yield_pct) if f.yield_pct else None,
            pe_ratio=float(f.pe_ratio) if f.pe_ratio else None,
            pb_ratio=float(f.pb_ratio) if f.pb_ratio else None,
            data_as_of_date=f.data_as_of_date,
            scores=ScoreOut(
                beta_score=float(scores.beta_score) if scores and scores.beta_score else None,
                r_squared_score=float(scores.r_squared_score) if scores and scores.r_squared_score else None,
                up_capture_score=float(scores.up_capture_score) if scores and scores.up_capture_score else None,
                down_capture_score=float(scores.down_capture_score) if scores and scores.down_capture_score else None,
                sharpe_score=float(scores.sharpe_score) if scores and scores.sharpe_score else None,
                tracking_error_score=float(scores.tracking_error_score) if scores and scores.tracking_error_score else None,
                sortino_score=float(scores.sortino_score) if scores and scores.sortino_score else None,
                treynor_score=float(scores.treynor_score) if scores and scores.treynor_score else None,
                info_ratio_score=float(scores.info_ratio_score) if scores and scores.info_ratio_score else None,
                kurtosis_score=float(scores.kurtosis_score) if scores and scores.kurtosis_score else None,
                drawdown_score=float(scores.drawdown_score) if scores and scores.drawdown_score else None,
                skewness_score=float(scores.skewness_score) if scores and scores.skewness_score else None,
                risk_score=float(scores.risk_score) if scores and scores.risk_score else None,
                alpha_score=float(scores.alpha_score) if scores and scores.alpha_score else None,
                yield_score=float(scores.yield_score) if scores and scores.yield_score else None,
                relative_return_score=float(scores.relative_return_score) if scores and scores.relative_return_score else None,
                price_score=float(scores.price_score) if scores and scores.price_score else None,
                fee_score=float(scores.fee_score) if scores and scores.fee_score else None,
                return_score=float(scores.return_score) if scores and scores.return_score else None,
                total_rr_score=float(scores.total_rr_score) if scores and scores.total_rr_score else None,
                market_cap_score=float(scores.market_cap_score) if scores and scores.market_cap_score else None,
                turnover_score=float(scores.turnover_score) if scores and scores.turnover_score else None,
                total_gpa_score=float(scores.total_gpa_score) if scores and scores.total_gpa_score else None,
                category_rank=scores.category_rank if scores else None,
                global_rank=scores.global_rank if scores else None,
            ) if scores else None
        )
        result.append(fund_out)

    return result


@router.get("/funds/{ticker}")
async def get_fund(ticker: str, db: Session = Depends(get_db)):
    """Get a single fund with all its data and scores."""
    fund = db.query(Fund).filter(Fund.ticker == ticker.upper()).first()
    if not fund:
        raise HTTPException(status_code=404, detail=f"Fund {ticker} not found")

    cat = db.query(FundCategory).filter(FundCategory.id == fund.category_id).first()
    scores = db.query(FundScore).filter(FundScore.fund_id == fund.id).first()

    result = {
        "id": fund.id,
        "ticker": fund.ticker,
        "name": fund.name,
        "fund_type": fund.fund_type,
        "category_id": fund.category_id,
        "category_name": cat.name if cat else None,
        "parent_category": cat.parent_category if cat else None,
        "fund_age_years": float(fund.fund_age_years) if fund.fund_age_years else None,
        "net_expense_ratio": float(fund.net_expense_ratio) if fund.net_expense_ratio else None,
        "turnover": float(fund.turnover) if fund.turnover else None,
        "market_cap": float(fund.market_cap) if fund.market_cap else None,
        "yield_pct": float(fund.yield_pct) if fund.yield_pct else None,
        "pe_ratio": float(fund.pe_ratio) if fund.pe_ratio else None,
        "pb_ratio": float(fund.pb_ratio) if fund.pb_ratio else None,
        "data_as_of_date": str(fund.data_as_of_date) if fund.data_as_of_date else None,
    }

    if scores:
        result["scores"] = {
            col.name: float(getattr(scores, col.name)) if getattr(scores, col.name) is not None else None
            for col in FundScore.__table__.columns
            if col.name not in ('id', 'fund_id', 'calculated_at')
        }

    return result


@router.put("/funds/{ticker}")
async def update_fund(ticker: str, fund_data: FundUpdate, db: Session = Depends(get_db)):
    """Update a single fund's data."""
    fund = db.query(Fund).filter(Fund.ticker == ticker.upper()).first()
    if not fund:
        raise HTTPException(status_code=404, detail=f"Fund {ticker} not found")

    update_data = fund_data.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        if hasattr(fund, key):
            setattr(fund, key, value)
    fund.updated_at = datetime.utcnow()

    db.commit()

    # Recalculate scores
    _recalculate_all(db)

    return {"success": True, "message": f"Fund {ticker} updated"}


@router.delete("/funds/{ticker}")
async def delete_fund(ticker: str, db: Session = Depends(get_db)):
    """Delete a fund."""
    fund = db.query(Fund).filter(Fund.ticker == ticker.upper()).first()
    if not fund:
        raise HTTPException(status_code=404, detail=f"Fund {ticker} not found")

    db.delete(fund)
    db.commit()

    # Re-rank remaining funds
    _recalculate_all(db)

    return {"success": True, "message": f"Fund {ticker} deleted"}


# =====================================================================
# UPLOAD
# =====================================================================

@router.post("/funds/upload", response_model=UploadSummary)
async def upload_funds(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Upload Excel/CSV file with fund data."""
    filename = file.filename or ""

    if not filename.endswith(('.xlsx', '.xls', '.csv')):
        raise HTTPException(status_code=400, detail="File must be Excel (.xlsx/.xls) or CSV format")

    content = await file.read()
    added = 0
    updated = 0
    errors = 0
    error_details = []

    try:
        rows = []
        headers = []

        if filename.endswith('.csv'):
            text = content.decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(text))
            headers = reader.fieldnames or []
            rows = list(reader)
        else:
            # Excel
            try:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
                ws = wb.active
                all_rows = list(ws.iter_rows(values_only=True))
                if len(all_rows) < 2:
                    raise HTTPException(status_code=400, detail="File must have a header row and at least one data row")
                headers = [str(h).strip() if h else "" for h in all_rows[0]]
                for row_vals in all_rows[1:]:
                    row_dict = {}
                    for i, h in enumerate(headers):
                        if i < len(row_vals):
                            row_dict[h] = row_vals[i]
                    rows.append(row_dict)
                wb.close()
            except ImportError:
                raise HTTPException(status_code=500, detail="openpyxl not installed. Use CSV format or install openpyxl.")

        # Build category name -> id mapping
        categories = db.query(FundCategory).all()
        cat_name_map = {c.name.lower(): c.id for c in categories}

        # Map headers to DB fields
        header_map = {}
        for h in headers:
            normalized = h.strip().lower()
            if normalized in COLUMN_MAP:
                header_map[h] = COLUMN_MAP[normalized]

        for row_idx, row in enumerate(rows):
            try:
                mapped = {}
                for orig_header, db_field in header_map.items():
                    val = row.get(orig_header)
                    if val is None or (isinstance(val, str) and val.strip() == ""):
                        mapped[db_field] = None
                    elif db_field in NUMERIC_FIELDS:
                        try:
                            mapped[db_field] = float(val)
                        except (ValueError, TypeError):
                            mapped[db_field] = None
                    elif db_field == "inception_date":
                        if isinstance(val, datetime):
                            mapped[db_field] = val.date()
                        elif isinstance(val, date):
                            mapped[db_field] = val
                        elif isinstance(val, str):
                            for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y"):
                                try:
                                    mapped[db_field] = datetime.strptime(val.strip(), fmt).date()
                                    break
                                except ValueError:
                                    continue
                    else:
                        mapped[db_field] = str(val).strip() if val else None

                ticker = mapped.get("ticker")
                if not ticker:
                    errors += 1
                    error_details.append(f"Row {row_idx + 2}: Missing ticker")
                    continue

                ticker = ticker.upper().strip()

                # Resolve category
                cat_name = mapped.pop("_category_name", None)
                cat_id = None
                if cat_name:
                    cat_id = cat_name_map.get(cat_name.lower())
                    if not cat_id:
                        # Try partial match
                        for cname, cid in cat_name_map.items():
                            if cat_name.lower() in cname or cname in cat_name.lower():
                                cat_id = cid
                                break

                # Check if fund exists (upsert)
                existing = db.query(Fund).filter(Fund.ticker == ticker).first()
                if existing:
                    for key, value in mapped.items():
                        if key not in ("ticker",) and hasattr(existing, key) and value is not None:
                            setattr(existing, key, value)
                    if cat_id:
                        existing.category_id = cat_id
                    existing.updated_at = datetime.utcnow()
                    updated += 1
                else:
                    fund = Fund(ticker=ticker, category_id=cat_id)
                    for key, value in mapped.items():
                        if key != "ticker" and key != "_category_name" and hasattr(fund, key):
                            setattr(fund, key, value)
                    db.add(fund)
                    added += 1

            except Exception as e:
                errors += 1
                error_details.append(f"Row {row_idx + 2}: {str(e)}")

        db.commit()

        # Recalculate all scores
        if added > 0 or updated > 0:
            _recalculate_all(db)

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to process file: {str(e)}")

    return UploadSummary(
        added=added,
        updated=updated,
        errors=errors,
        error_details=error_details[:20]  # Limit error details
    )


# =====================================================================
# RANKINGS / SCORES
# =====================================================================

@router.get("/scores")
async def get_ranked_funds(
    category: Optional[int] = None,
    sort: str = "gpa_score",
    order: str = "desc",
    page: int = 1,
    limit: int = 25,
    search: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Get ranked funds with optional category filter."""
    query = db.query(Fund, FundScore, FundCategory).outerjoin(
        FundScore, Fund.id == FundScore.fund_id
    ).outerjoin(
        FundCategory, Fund.category_id == FundCategory.id
    )

    if category:
        query = query.filter(Fund.category_id == category)
    if search:
        search_term = f"%{search}%"
        query = query.filter(
            (Fund.ticker.ilike(search_term)) | (Fund.name.ilike(search_term))
        )

    # Sort
    sort_map = {
        "gpa_score": FundScore.total_gpa_score,
        "risk_score": FundScore.risk_score,
        "return_score": FundScore.return_score,
        "rr_score": FundScore.total_rr_score,
        "ticker": Fund.ticker,
        "name": Fund.name,
        "category_rank": FundScore.category_rank,
        "global_rank": FundScore.global_rank,
    }
    sort_col = sort_map.get(sort, FundScore.total_gpa_score)
    if sort_col is not None:
        if order == "asc":
            query = query.order_by(asc(sort_col))
        else:
            query = query.order_by(desc(sort_col).nulls_last())
    else:
        query = query.order_by(desc(FundScore.total_gpa_score).nulls_last())

    total = query.count()
    results = query.offset((page - 1) * limit).limit(limit).all()

    # Get data_as_of
    latest_date = db.query(func.max(Fund.data_as_of_date)).scalar()

    funds_out = []
    for idx, (fund, score, cat) in enumerate(results):
        rank = ((page - 1) * limit) + idx + 1
        if category and score and score.category_rank:
            rank = score.category_rank
        elif score and score.global_rank:
            rank = score.global_rank

        fund_out = {
            "rank": rank,
            "ticker": fund.ticker,
            "name": fund.name,
            "fund_type": fund.fund_type,
            "category_name": cat.name if cat else None,
            "parent_category": cat.parent_category if cat else None,
            "category_id": fund.category_id,
            "fund_age_years": float(fund.fund_age_years) if fund.fund_age_years else None,
            "net_expense_ratio": float(fund.net_expense_ratio) if fund.net_expense_ratio else None,
            "turnover": float(fund.turnover) if fund.turnover else None,
            "market_cap": float(fund.market_cap) if fund.market_cap else None,
            "yield_pct": float(fund.yield_pct) if fund.yield_pct else None,
            "pe_ratio": float(fund.pe_ratio) if fund.pe_ratio else None,
            "pb_ratio": float(fund.pb_ratio) if fund.pb_ratio else None,
            "total_gpa_score": float(score.total_gpa_score) if score and score.total_gpa_score else None,
            "risk_score": float(score.risk_score) if score and score.risk_score else None,
            "return_score": float(score.return_score) if score and score.return_score else None,
            "total_rr_score": float(score.total_rr_score) if score and score.total_rr_score else None,
            "category_rank": score.category_rank if score else None,
            "global_rank": score.global_rank if score else None,
            "scores": None,
        }

        if score:
            fund_out["scores"] = {
                col.name: float(getattr(score, col.name)) if getattr(score, col.name) is not None else None
                for col in FundScore.__table__.columns
                if col.name not in ('id', 'fund_id', 'calculated_at')
            }

        funds_out.append(fund_out)

    # Category summary
    category_summary = None
    if category:
        cat_obj = db.query(FundCategory).filter(FundCategory.id == category).first()
        if cat_obj:
            cat_funds = db.query(FundScore).join(Fund).filter(Fund.category_id == category).all()
            if cat_funds:
                gpas = [float(s.total_gpa_score) for s in cat_funds if s.total_gpa_score is not None]
                risk_scores = [float(s.risk_score) for s in cat_funds if s.risk_score is not None]
                expenses = db.query(Fund.net_expense_ratio).filter(
                    Fund.category_id == category,
                    Fund.net_expense_ratio.isnot(None)
                ).all()
                exp_vals = [float(e[0]) for e in expenses if e[0] is not None]

                # Find highest GPA ticker
                best = db.query(Fund, FundScore).join(FundScore).filter(
                    Fund.category_id == category
                ).order_by(desc(FundScore.total_gpa_score)).first()

                category_summary = {
                    "category_name": cat_obj.name,
                    "fund_count": len(cat_funds),
                    "avg_gpa": round(sum(gpas) / len(gpas), 2) if gpas else None,
                    "highest_gpa": round(max(gpas), 2) if gpas else None,
                    "highest_gpa_ticker": best[0].ticker if best else None,
                    "lowest_gpa": round(min(gpas), 2) if gpas else None,
                    "avg_expense": round(sum(exp_vals) / len(exp_vals), 4) if exp_vals else None,
                    "avg_risk_score": round(sum(risk_scores) / len(risk_scores), 2) if risk_scores else None,
                }

    return {
        "funds": funds_out,
        "total": total,
        "page": page,
        "limit": limit,
        "category": category,
        "data_as_of": str(latest_date) if latest_date else None,
        "category_summary": category_summary,
    }


@router.get("/scores/{ticker}/detail")
async def get_fund_score_detail(ticker: str, db: Session = Depends(get_db)):
    """Get full score breakdown for a single fund."""
    fund = db.query(Fund).filter(Fund.ticker == ticker.upper()).first()
    if not fund:
        raise HTTPException(status_code=404, detail=f"Fund {ticker} not found")

    score = db.query(FundScore).filter(FundScore.fund_id == fund.id).first()
    cat = db.query(FundCategory).filter(FundCategory.id == fund.category_id).first()

    result = {
        "ticker": fund.ticker,
        "name": fund.name,
        "fund_type": fund.fund_type,
        "category_name": cat.name if cat else None,
        "parent_category": cat.parent_category if cat else None,
        "fund_age_years": float(fund.fund_age_years) if fund.fund_age_years else None,
        "net_expense_ratio": float(fund.net_expense_ratio) if fund.net_expense_ratio else None,
        "turnover": float(fund.turnover) if fund.turnover else None,
        "market_cap": float(fund.market_cap) if fund.market_cap else None,
        "yield_pct": float(fund.yield_pct) if fund.yield_pct else None,
        "pe_ratio": float(fund.pe_ratio) if fund.pe_ratio else None,
        "pb_ratio": float(fund.pb_ratio) if fund.pb_ratio else None,
    }

    if score:
        result["scores"] = {
            col.name: float(getattr(score, col.name)) if getattr(score, col.name) is not None else None
            for col in FundScore.__table__.columns
            if col.name not in ('id', 'fund_id', 'calculated_at')
        }
    else:
        result["scores"] = None

    return result


# =====================================================================
# RECALCULATE
# =====================================================================

@router.post("/recalculate", response_model=RecalculateResponse)
async def recalculate_rankings(db: Session = Depends(get_db)):
    """Recalculate all fund scores and rankings."""
    count = _recalculate_all(db)
    return RecalculateResponse(
        success=True,
        funds_calculated=count,
        message=f"Successfully recalculated scores for {count} funds"
    )


# =====================================================================
# EXPORT
# =====================================================================

@router.get("/export")
async def export_rankings(
    category: Optional[int] = None,
    format: str = "csv",
    db: Session = Depends(get_db)
):
    """Export rankings as CSV or Excel."""
    query = db.query(Fund, FundScore, FundCategory).outerjoin(
        FundScore, Fund.id == FundScore.fund_id
    ).outerjoin(
        FundCategory, Fund.category_id == FundCategory.id
    ).order_by(desc(FundScore.total_gpa_score).nulls_last())

    if category:
        query = query.filter(Fund.category_id == category)

    results = query.all()

    headers = [
        "Rank", "Ticker", "Name", "Type", "Category", "GPA Score",
        "Risk Score", "Return Score", "RR Score",
        "Beta Score", "R² Score", "Up Capture Score", "Down Capture Score",
        "Sharpe Score", "Tracking Error Score", "Sortino Score", "Treynor Score",
        "Info Ratio Score", "Kurtosis Score", "Drawdown Score", "Skewness Score",
        "Alpha Score", "Yield Score", "Relative Return Score", "Price Score", "Fee Score",
        "Market Cap Score", "Turnover Score",
        "Category Rank", "Global Rank",
        "Expense Ratio", "Turnover", "Market Cap (M)", "Yield", "P/E", "P/B",
        "Fund Age (Yrs)"
    ]

    rows = []
    for idx, (fund, score, cat) in enumerate(results):
        rank = score.global_rank if score and score.global_rank else idx + 1
        if category and score and score.category_rank:
            rank = score.category_rank

        row = [
            rank,
            fund.ticker,
            fund.name,
            fund.fund_type,
            cat.name if cat else "",
            f"{float(score.total_gpa_score):.2f}" if score and score.total_gpa_score else "",
            f"{float(score.risk_score):.2f}" if score and score.risk_score else "",
            f"{float(score.return_score):.2f}" if score and score.return_score else "",
            f"{float(score.total_rr_score):.2f}" if score and score.total_rr_score else "",
            f"{float(score.beta_score):.2f}" if score and score.beta_score else "",
            f"{float(score.r_squared_score):.2f}" if score and score.r_squared_score else "",
            f"{float(score.up_capture_score):.2f}" if score and score.up_capture_score else "",
            f"{float(score.down_capture_score):.2f}" if score and score.down_capture_score else "",
            f"{float(score.sharpe_score):.2f}" if score and score.sharpe_score else "",
            f"{float(score.tracking_error_score):.2f}" if score and score.tracking_error_score else "",
            f"{float(score.sortino_score):.2f}" if score and score.sortino_score else "",
            f"{float(score.treynor_score):.2f}" if score and score.treynor_score else "",
            f"{float(score.info_ratio_score):.2f}" if score and score.info_ratio_score else "",
            f"{float(score.kurtosis_score):.2f}" if score and score.kurtosis_score else "",
            f"{float(score.drawdown_score):.2f}" if score and score.drawdown_score else "",
            f"{float(score.skewness_score):.2f}" if score and score.skewness_score else "",
            f"{float(score.alpha_score):.2f}" if score and score.alpha_score else "",
            f"{float(score.yield_score):.2f}" if score and score.yield_score else "",
            f"{float(score.relative_return_score):.4f}" if score and score.relative_return_score else "",
            f"{float(score.price_score):.2f}" if score and score.price_score else "",
            f"{float(score.fee_score):.2f}" if score and score.fee_score else "",
            f"{float(score.market_cap_score):.2f}" if score and score.market_cap_score else "",
            f"{float(score.turnover_score):.2f}" if score and score.turnover_score else "",
            score.category_rank if score else "",
            score.global_rank if score else "",
            f"{float(fund.net_expense_ratio) * 100:.2f}%" if fund.net_expense_ratio else "",
            f"{float(fund.turnover) * 100:.0f}%" if fund.turnover else "",
            f"{float(fund.market_cap):.0f}" if fund.market_cap else "",
            f"{float(fund.yield_pct) * 100:.2f}%" if fund.yield_pct else "",
            f"{float(fund.pe_ratio):.1f}" if fund.pe_ratio else "",
            f"{float(fund.pb_ratio):.2f}" if fund.pb_ratio else "",
            f"{float(fund.fund_age_years):.1f}" if fund.fund_age_years else "",
        ]
        rows.append(row)

    if format == "xlsx":
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Fund Rankings"
            ws.append(headers)
            for row in rows:
                ws.append(row)

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=fund_rankings.xlsx"}
            )
        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl required for Excel export")
    else:
        # CSV
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(rows)

        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=fund_rankings.csv"}
        )


# =====================================================================
# TEMPLATE
# =====================================================================

@router.get("/template")
async def download_template():
    """Download a blank upload template as CSV."""
    template_headers = [
        "Ticker", "Name", "Fund Type", "Category",
        "Inception Date", "Fund Age (Years)", "Net Expense Ratio", "Turnover",
        "Market Cap (M)", "Yield", "P/E Ratio", "P/B Ratio",
        "Beta 3yr", "Beta 5yr",
        "R-Squared 3yr", "R-Squared 5yr",
        "Up Capture 3yr", "Up Capture 5yr",
        "Down Capture 3yr", "Down Capture 5yr",
        "Sharpe 3yr", "Sharpe 5yr",
        "Tracking Error 3yr", "Tracking Error 5yr",
        "Sortino 3yr", "Sortino 5yr",
        "Treynor 3yr", "Treynor 5yr",
        "Info Ratio 3yr", "Info Ratio 5yr",
        "Kurtosis 3yr", "Kurtosis 5yr",
        "Max Drawdown 3yr", "Max Drawdown 5yr",
        "Skewness 3yr", "Skewness 5yr",
        "Alpha 3yr", "Alpha 5yr",
        "Return QTD", "Return YTD", "Return 1yr",
        "Return 3yr", "Return 5yr", "Return 10yr",
        "BM Return 1yr", "BM Return 3yr", "BM Return 5yr", "BM Return 10yr",
        "Batting Avg 3yr", "Batting Avg 5yr"
    ]

    # Example rows
    example_rows = [
        [
            "SPY", "SPDR S&P 500 ETF", "ETF", "Large Cap Blend",
            "1993-01-22", "32", "0.0009", "0.02",
            "1500", "0.013", "22.5", "4.2",
            "1.00", "1.00",
            "99.5", "99.3",
            "1.00", "1.00",
            "1.00", "1.00",
            "0.85", "0.90",
            "1.2", "1.1",
            "1.10", "1.15",
            "0.10", "0.10",
            "0.05", "0.08",
            "0.8", "1.0",
            "15.2", "18.5",
            "-0.3", "-0.25",
            "0.5", "0.3",
            "0.02", "0.12", "0.18",
            "0.10", "0.11", "0.12",
            "0.18", "0.10", "0.11", "0.12",
            "0.50", "0.48"
        ],
        [
            "AGG", "iShares Core U.S. Aggregate Bond ETF", "ETF", "US Aggregate Bond",
            "2003-09-22", "22", "0.0003", "0.05",
            "300", "0.035", "", "",
            "0.05", "0.04",
            "50.2", "48.5",
            "0.20", "0.18",
            "0.15", "0.12",
            "0.30", "0.35",
            "2.5", "2.3",
            "0.40", "0.45",
            "0.03", "0.03",
            "0.10", "0.12",
            "0.5", "0.7",
            "5.2", "8.1",
            "-0.2", "-0.18",
            "0.8", "0.6",
            "0.01", "0.03", "0.04",
            "0.02", "0.03", "",
            "0.04", "0.02", "0.03", "",
            "0.40", "0.38"
        ]
    ]

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(template_headers)
    for row in example_rows:
        writer.writerow(row)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=fund_upload_template.csv"}
    )
